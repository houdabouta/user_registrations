import openpyxl
import argparse
from datetime import datetime

def filter_data(input_file, output_file, start_date, end_date, location):
    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook.active

    filtered_workbook = openpyxl.Workbook()
    filtered_sheet = filtered_workbook.active
    filtered_sheet.title = "Filtered Data"

    # Copy header
    for cell in sheet[1]:
        filtered_sheet.cell(row=1, column=cell.column, value=cell.value)

    # Find the column indexes for authDate and location
    headers = {cell.value: cell.column for cell in sheet[1]}
    auth_date_col_idx = headers.get('authDate')
    location_col_idx = headers.get('location')

    start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
    end_date = datetime.strptime(end_date, "%Y-%m-%d").date()

    row_index = 2
    for row in sheet.iter_rows(min_row=2, values_only=True):
        auth_date_str = row[auth_date_col_idx - 1]
        item_location = row[location_col_idx - 1]

        if auth_date_str:
            auth_date = datetime.strptime(auth_date_str, "%Y-%m-%dT%H:%M:%S%z").date()
            if start_date <= auth_date <= end_date and str(item_location) == location:
                for col_index, cell_value in enumerate(row, start=1):
                    filtered_sheet.cell(row=row_index, column=col_index, value=cell_value)
                row_index += 1

    filtered_workbook.save(output_file)

def main():
    parser = argparse.ArgumentParser(description="Filter history data Excel file.")
    parser.add_argument("--input_file", type=str, required=True, help="Path to the input Excel file")
    parser.add_argument("--output_file", type=str, required=True, help="Path to the output Excel file")
    parser.add_argument("--start_date", type=str, required=True, help="Start date in YYYY-MM-DD format")
    parser.add_argument("--end_date", type=str, required=True, help="End date in YYYY-MM-DD format")
    parser.add_argument("--location", type=str, required=True, help="Location to filter by")

    args = parser.parse_args()

    filter_data(args.input_file, args.output_file, args.start_date, args.end_date, args.location)

if __name__ == "__main__":
    main()
